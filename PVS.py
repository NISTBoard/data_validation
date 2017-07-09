# Author: Ron Jones
# Date Created: 7-3-17
# Date Last Modified: 7-4-17
# Purpose: Check Privacy Overlay Excel Sheet with Master Data Sheet
# Status: Working perfectly with MDS and CDS_Overlay_Final2.xlsx as of July 4, 2017

'''Note: The "compare dicts function iterates through every
correct combination of entries from the overlay and data files to check
for any discrepancies, then checks every entry from the overlay against
the data to see if there are any entire records erroneously absent from
the MDS'''


# Import openpyxl module to allow python to access data from Excel documents
import openpyxl as xl, sys

def main():
    # Pull data from workbooks
    data = xl.load_workbook(sys.argv[1])
    overlay = xl.load_workbook(sys.argv[2])


    # Pull worksheets from workbooks
    data_sheet = data.get_sheet_by_name('Data')
    overlay_sheet = overlay.get_sheet_by_name('Table 1')


    # Open output file (validation comments) for writing
    comments = open('Privacy_Validation_Comments', 'w')
    #Write heading to output file
    comments.write("Inconsistencies:" + "\n" + "\n" + "Note: These could be do to an issue with the overlay or MDS" + "\n" + "\n")

    # Open empty dictionary for overlay info
    overlay_dict = {}

    # Open empty dictionary for master info
    data_dict = {}

    populate_overlay_dict(overlay_sheet, overlay_dict)

    populate_data_dict(data_sheet, data_dict)

    compare_dicts(data_dict, overlay_dict, comments)

def populate_overlay_dict(sheet, inp_dict):
    for i in range(5, 236):
        #Initialize list with Transfer value, then append Access and Multilevel values (Order: Transfer, Access, Multilevel)
        inp_dict[sheet.cell(row=i, column=1).value] = [sheet.cell(row=i, column=2).value]
        inp_dict[sheet.cell(row=i, column=1).value].append(sheet.cell(row=i, column=7).value)
        inp_dict[sheet.cell(row=i, column=1).value].append(sheet.cell(row=i, column=10).value)
        inp_dict[sheet.cell(row=i, column=1).value].append(sheet.cell(row=i, column=11).value)
    #print("Overlay Dictionary: ", inp_dict)

def populate_data_dict(worksheet, inp):
    for i in range(4, worksheet.max_row + 1):
        inp[worksheet.cell(row=i, column=3).value] = [worksheet.cell(row=i, column=58).value]
        inp[worksheet.cell(row=i, column=3).value].append(worksheet.cell(row=i, column=59).value)
        inp[worksheet.cell(row=i, column=3).value].append(worksheet.cell(row=i, column=60).value)
        inp[worksheet.cell(row=i, column=3).value].append(worksheet.cell(row=i, column=61).value)
    #print("Data Dict: ", inp)


# Function to compare overlay and data dictionaries: working!!
def compare_dicts(data, overlay, outfile):
    # Initiate error message switch
    switch = 0
    # Loop through all control names in data file to find errors
    for key in data:
        # Compare to each control name in overlay file
        for key2 in overlay:
            if key == key2:
                # Ensure every entry in data object matches corresponding entry in overlay object
                for i in range(len(data[key])):
                    if data[key][i] == overlay[key2][i]:
                        continue
                    # If inconsistency found, write to output file and flip error switch
                    else:
                        outfile.write("Inconsistency found in requirement " + str(key) + "\n" + "\n")
                        switch = 1
                        break

    # For loop to check for missing records in data file
    for key2 in overlay:
        # If a record required from the overlay is missing from the data, write error message to the output file and flip switch
        if not key2 in data and key2 != None:
            outfile.write("MDS is missing the entire record for requirement " + str(key2) + " from the overlay" + "\n")
            switch = 1

    for key in data:
        if not key in overlay:
            for elt in data[key]:
                if elt != None:
                    outfile.write("Erroneous addition to MDS in requirement " + str(key) + "\n" + "\n")
                    switch = 1
                    break
        continue

    # Print message based on switch position
    if switch == 0:
        print("\n" + "No discrepancies found")

    else:
        print("\n" + "There were some discrepancies. Check 'Privacy_Validation_Comments' for more information.")



main()
