# Import openpyxl module to allow python to access data from Excel documents
import openpyxl as xl

def main():
    # Pull data from workbooks
    data = xl.load_workbook('FM_Data.xlsx')
    overlay = xl.load_workbook('FM_Overlay.xlsx')


    # Pull worksheets from workbooks
    data_sheet = data.get_sheet_by_name('Scoped NIST Controls')
    overlay_sheet = overlay.get_sheet_by_name('FM Overlay-AC')


    # Open output file (validation comments) for writing
    comments = open('FM Validation Comments', 'w')
    #Write heading to output file
    comments.write("Missing Requirements:" + "\n" + "\n" + "Note: error could be caused by a typo in either input file" + "\n" + "\n")
    
    
    # Open empty requirements dictionary for data from "FM_Data.xlsx"
    # Dictionary key = name, like AC-1, value = list of every line-separated value from column C
    reqs_dict = {}


    # Open empty data dictionary for data from "FM_Overlay.xlsx"
    # Dictionary: key = name, if in the dictionary already, append value to existing value list, else add the entry and initiate value list
    data_dict = {}


    # Open empty dictionary to store control ranks from "FM_Data.xlsx", column D
    rank_master_dict = {}


    # Open empty dictionary to store control ranks from "FM_Overlay.xlsx", column D
    rank_data_dict = {}
    

    # Populate requirements and data dictionaries and compare
    populate_req_dict(data_sheet, reqs_dict)
    populate_data_dict(overlay_sheet, data_dict)
    compare_reqs(reqs_dict, data_dict, comments)

    # Write label heading into output file for objective 2: ensuring label consistency
    comments.write("\n" + "Incorrect Labels:" + "\n" + "\n")



    # Populate label dictionaries (rank_master_dict input comes from "FM_Data.xlsx", column 4 and rank_data_dict
    # input comes from "FM_Overlay.xlsx", column 4) and compare using the compare_ranks function
    pop_master_rank_reqs(data_sheet, rank_master_dict)
    pop_data_rank(overlay_sheet, rank_data_dict)
    compare_ranks(rank_master_dict, rank_data_dict, comments)


    # Close output file
    comments.close()


# Function to populate requirement dictionary: working!!
def populate_req_dict(worksheet, req_dict):
    # Iterate through every row in "FM_Data.xlsx"
    for i in range(2,worksheet.max_row + 1):#25, or 2, max
        # Split up values in column 3, and make them into a list
        reqs_lst = worksheet.cell(row=i, column=3).value.split('\n')

        # Create a new entry in the req_dict dictionary (key = name, like AC-1, value = newly created list from column 3 of the row)
        req_dict[worksheet.cell(row=i, column=1).value] = reqs_lst
        
    #Can uncomment the line below to see the result of the dictionary population
    #print("Req dict:", req_dict)
        

    # Data cleaning: the code below removes blanks from dictionaries due to end of cells in spreadsheets (there are many empty spaces where
    # people hit "Enter" in the excel cells but didn't write anything, so this step should make the program more resilient to human input error)
        
    # Iterate through keys in requirements dictionary
    for key in req_dict:
        # Iterate through corresponding list of each key in requirement dictionary
        for elt in req_dict[key]:
            # Strip any spaces, takes care of weird blanks
            elt = elt.strip()
            #Check if any stripped elements are now blank (call Ron for more details)
            if elt == "":
                req_dict[key].remove(elt)
                
    #Could print requirement dictionary below here for quality check of data cleaning
    #print("Requirement Dictionary post-clean: ", req_dict)



# Note: this program does not know how to switch between sheets in an excel workbook.
# Because of this, I simply copied all data from other sheets into sheet "FM Overlay-AC" "FM_Overlay.xlsx".
# That is also why that sheet has nearly 4000 records

# Function to populate data dictionary: working!
def populate_data_dict(worksheet, di):
    
    # Iterate through every row in "FM_Overlay.xlsx"
    for i in range(2,worksheet.max_row + 1):
        # If the name (column 1) is already in the dictionary, append the corresponding value (in column 3) to the existing list
        if worksheet.cell(row=i, column=1).value in di:
            di[worksheet.cell(row=i, column=1).value].append(worksheet.cell(row=i, column=3).value)

        # If the name (column 1) is not already in the dictionary, create a new entry in the dictionary and begin a new value list with column 3 value
        else:
            di[worksheet.cell(row=i, column=1).value] = [worksheet.cell(row=i, column=3).value]
            
    #Could uncomment the line below to see the result of the dictionary population
    #print("Data Dictionary: ", di)



# Compare requirements dictionary contents to data dictionaries: APPEARS TO BE WORKING!! (Need to test more output to make sure)
def compare_reqs(reqs, data, comment_file):
    # List to hold phrases program should ignore, as discussed in Matrix meeting. Can be edited as new info is discovered
    # about what these phrases mean
    phrases_to_ignore = ["FISCAM's CM-4 Series", "FISCAM's CM-4 series","CM-1 Disscussion","FISCAM's BP-3 series","CM-1 Disscussion","","Specific policies and procedures should be created to the meet the Requirement/Control Technique of this NIST Control","CM-1 Disscussion from SA-8"]

    # Counter to determine whether or not to print error message
    counter = 0


    # The for loop compares requirements from the requirements dictionary (data from "FM_Data.xlsx", called "reqs" here) to
    # requirements included in "FM_Overlay.xlsx (called "data" here) to see if any requirements are missing from "FM_Overlay.xlsx"
    # OR if there are any typos in "FM_Overlay.xlsx"

    
    # Iterate through every name (called "key") in requirements dictionary
    for key in reqs:
        # Compare every name (key) in requirements dictionary to every name in data dictionary (called "key2")
        for key2 in data:

            # Conditionals to determine if two records should be compared at all (call Ron for more details)
            if (key == key2) or (key == key2[0:4] and key2[len(key2) - 1] == ")" and (len(key2) == 7 or len(key2) == 8) and key2[4] == "(") or (key == key2[0:4] and len(key2) == 4) or (len(key2) >= 5 and key == key2[0:5]):
                
                # If a condition is met, iterate through the requirements list associated with the related requirements name (key, from "FM_Data.xlsx")
                for req in reqs[key]:

                    # If the requirement in the requirements list is also present in "FM_Overlay.xlsx", then move on
                    if req in data[key2]:
                        # Note to self: Could add a placeholder message here to visually show progress
                        continue
                    
                    # If the requirement is missing from "FM_Overlay.xlsx" and not in phrases to ignore, write
                    # an error to the output file, and turn on switch to write error message to the output file
                    else:
                        if not req in phrases_to_ignore:
                            counter = 1
                            comment_file.write("Requirement " + str(req) + " from " + str(key) + " is missing from " + str(key2) + "\n")
                            
            # If the name (key2) in the data dictionary does not match the name (key) in the requirement dictionary, move to the next comparison
                continue


    
    # Write title to output file, signaling that erroneously included requirements will now be addressed
    # These could be caused either because someone included an unnecessary requirement OR because there is a typo in
    # the requirement in "FM_Data.xlsx"
    comment_file.write("\n" + "Erroneously included requirements: " + "\n" + "Note: This could be caused by a typo in either input file" + "\n" + "\n")



    # The for loop below compares requirements in the opposite direction to see if any extra requirements
    # were erroneously included in "FM_Overlay.xlsx" or if there is a typo in "FM_Data.xlsx"

    # Iterate through names in data dictionary (from "FM_Overlay.xlsx")
    for key2 in data:
        
        # Compare every name in data dictionary (called "key 2") to every name in requirements dictionary (called "key")
        for key in reqs:

            # Same conditionals as in previous for loop to see if two names should be compared in the first place
            if (key == key2) or (key == key2[0:4] and key2[len(key2) - 1] == ")" and (len(key2) == 7 or len(key2) == 8) and key2[4] == "(") or (key == key2[0:4] and len(key2) == 4) or (len(key2) >= 5 and key == key2[0:5]):
                # If two records should be compared, iterate through the requirements list associated with related data name
                # (called "key2", from "FM_Overlay.xlsx")
                for elt in data[key2]:

                    # If the requirement found in the data is supposed to be there, move on
                    if elt in reqs[key]:
                        continue

                    # If the requirement found in the data is not supposed to be there and is not
                    # in "phrases_to_ignore", then turn on error switch (called "counter" here) and write error
                    # message to output file
                    else:
                        if not elt in phrases_to_ignore:
                            counter = 1
                            comment_file.write("Requirement " + str(elt) + " was erroneously included in " + str(key2) + "\n" + "\n")

            # If two names should not be compared in the first place, then move on to the next possible combo
            else:
                continue

            
    # If error switch (counter) is flipped, print error message
    if counter == 1:
        print("\n" + "There were some discrepancies. Check 'FM Validation Comments' for more details" + "\n")

    # If no discrepancies found and switch remains unflipped, print success message
    else:
        print("All requirements are accounted for")



# Function to create lookup dictionary for Control Rankings from requirements sheet: working!!!
def pop_master_rank_reqs(sheet, rank_dict):

    # Iterate through every row of "FM_Data.xlsx" and create a new rank_dict entry (key = name in col 1, value = label in col 4)
    for i in range(2,sheet.max_row + 1):#25 for test set
        rank_dict[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=4).value

        
    #Could uncomment below code for quality check on rank_dict output
    #print("Master Rank Dict:", rank_dict)



# Function to create Control Rankings dictionary from the actual data: working!!!
def pop_data_rank(input_sheet, ranking_dict):

    # Iterate through every row of "FM_Overlay.xlsx"
    for i in range(2,input_sheet.max_row + 1):#1488 for test set

        # If the name (column 1) is already in the dictionary, then append the corresponding label (column 4) to the existing label list
        if input_sheet.cell(row=i, column=1).value in ranking_dict:
            ranking_dict[input_sheet.cell(row=i, column=1).value].append(input_sheet.cell(row=i, column=4).value)

        # If the name (column 1) is not already in the dictionary, create a new entry (key = name from column 1, value = new list containing column 4 label)
        else:
            ranking_dict[input_sheet.cell(row=i, column=1).value] = [input_sheet.cell(row=i, column=4).value]

            
    #Could uncomment below code for quality check on ranking_dict output
    #print("Data Rank Dict:", ranking_dict)



# Function to compare Control Rankings Lookup dictionary to Control Rankings dictionary from the actual data: working!!!
def compare_ranks(master, data, comment_file):
    # Switch to determine whether or not to print error message
    switch = 0

    # List to contain unique names (call Ron for more detail)
    unique_rank_lst = []

    # Iterate through every name (key) in master lookup dictionary (made from "FM_Data.xlsx")
    for key in master:

        # For every name in master lookup, check against every name (key2) in data dictionary (made from "FM_Overlay.xlsx")
        # to see if they should be compared based on the conditional below
        for key2 in data:
            if (key == key2) or (key == key2[0:4] and key2[len(key2) - 1] == ")" and (len(key2) == 7 or len(key2) == 8) and key2[4] == "(") or (key == key2[0:4] and len(key2) == 4) or (len(key2) >= 5 and key == key2[0:5]):

                # If two names should be compared, iterate through attached list of labels from data dictionary (made from "FM_Overlay.xlsx")
                for elt in data[key2]:

                    # If the label is consistent with the master, move on
                    if elt == master[key]:
                        # Could add a placeholder message here to visually show progress
                        continue

                    # If the label is not consistent with the master, flip error switch
                    else:
                        switch = 1

                        # If label not consistent with the master and the error name in question is unique, append to unique list and
                        # write error to output file
                        if not key2 in unique_rank_lst:
                            unique_rank_lst.append(key2)
                            #FIGURE OUT HOW TO GIVE MESSAGE THAT IDENTIFIES RECORD
                            comment_file.write("Incorrect labels in " + str(key2) + "\n")


            # If two names cannot be compared, move on to the next combination
            else:
                continue

    # If error switch is flipped, print error message
    if switch == 1:
        print("Incorrect labels found. Check 'FM Validation Comments' for more details")

    # If error switch is not flipped, print success message and write success message to output file
    else:
        print("All labels correct")
        comment_file.write("All labels correct")



# Call to main function        
main()
