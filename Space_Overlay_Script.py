# Author: Ron Jones
# Date Created: 6-30-17
# Date Last Modified: 7-3-17
# Purpose: Check Space Overlay Excel Sheet with Master Data Sheet


# Import openpyxl module to allow python to access data from Excel documents
import openpyxl as xl, sys

def main():
    # Pull data from workbooks
    data = xl.load_workbook(sys.argv[1])
    overlay = xl.load_workbook(sys.argv[2])


    # Pull worksheets from workbooks
    data_sheet = data.get_sheet_by_name('Data')
    overlay_sheet = overlay.get_sheet_by_name('Table 1')


    # Open output file (validation comments) for writing
    comments = open('Space Overlay Comments', 'w')
    #Write heading to output file
    comments.write("Inconsistencies:" + "\n" + "\n")

    # Open empty dictionary for overlay info
    overlay_dict = {}

    # Open empty dictionary for master info
    data_dict = {}
    
    populate_overlay_dict(overlay_sheet, overlay_dict)

    populate_data_dict(data_sheet, data_dict)

    compare_dicts(data_dict, overlay_dict, comments)


# Function to populate overlay dictionary: working!!
def populate_overlay_dict(sheet, inp_dict):
    titles = ['CONTROL', 'SPACE PLATFORM OVERLAY']
    for i in range(67, 186): #67, 186 = actual, #80, 81 = test
        if not sheet.cell(row=i, column=1).value in titles: 
            inp_dict[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=2).value
    print("Overlay dictionary: ", inp_dict)


# Function to populate data dictionary: working!! 
def populate_data_dict(worksheet, inp):
    for i in range(4, worksheet.max_row + 1): #4, worksheet.max_row + 1 = actual, #339, 344 = test
        if not worksheet.cell(row=i, column=3).value in inp:
            inp[worksheet.cell(row=i, column=3).value] = [worksheet.cell(row=i, column=62).value]
        else:
            inp[worksheet.cell(row=i, column=3).value].append(worksheet.cell(row=i, column=62).value)
    print("Data Dict: ", inp)


# Function to compare overlay and data dictionaries: working!!
def compare_dicts(data, overlay, outfile):
    switch = 0

    #For loop to check for incorrect/missing entries
    for key in data:
        for key2 in overlay:
            if key == key2:
                for elt in data[key]:
                    if elt == overlay[key2]:
                        #Can uncomment for visual evidence that loop executed
                        #print("Data validated "  + str(key) + "    " + str(key2))
                        continue
                    else:
                        outfile.write("Discrepancy with control " + str(key) + "\n" + "\n")
                        switch = 1
                        break
            continue

    #For loop to check for missing records
    for key2 in overlay:
        if not key2 in data:
            outfile.write(((str(key2) + " should include a " + str(overlay[key2]) + " in the overlay column of MDS, but the record itself does not exist" + "\n" + "\n")))
            switch = 1
            
    if switch == 0:
        print("No discrepancies found")
    else:
        print("There were some discrepancies. Check 'Space_Overlay_Comments for more information")
        

#Call to main function
main()
